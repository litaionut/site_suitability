"""
Parsers for different input formats to site-package-v1 JSON.
"""
import json
import csv
import io
import zipfile
from typing import Dict, List, Any, Optional, Tuple
from bs4 import BeautifulSoup
import openpyxl


class ParseError(Exception):
    """Exception raised when parsing fails."""
    pass


class BaseParser:
    """Base parser class."""

    def parse(self, file_obj) -> Dict[str, Any]:
        """Parse file and return site-package-v1 dict."""
        raise NotImplementedError


class JSONParser(BaseParser):
    """Parse JSON site-package-v1 format."""

    def parse(self, file_obj) -> Dict[str, Any]:
        """Parse JSON file."""
        try:
            data = json.load(file_obj)
            if not isinstance(data, dict):
                raise ParseError("JSON must be an object (dict)")
            return data
        except json.JSONDecodeError as e:
            raise ParseError(f"Invalid JSON: {e}")


class ExcelParser(BaseParser):
    """Parse Excel .xlsx files."""

    SHEET_ALIASES = {
        'site': ['Site', 'SITE', 'site'],
        'layout': ['Layout', 'LAYOUT', 'layout', 'WTGs', 'wtgs', 'Turbines', 'turbines', 'TURBINES'],
        'wtg_models': ['WtgModels', 'WTGMODELS', 'wtgmodels', 'Models', 'MODELS', 'models'],
        'power_curve': ['PowerCurve', 'POWERCURVE', 'powercurve', 'Power', 'POWER', 'power'],
        'ct_curve': ['CtCurve', 'CTCURVE', 'ctcurve', 'Ct', 'CT', 'ct'],
        'hub_climate': ['HubClimate', 'HUBCLIMATE', 'hubclimate', 'Climate', 'CLIMATE', 'climate'],
        'ti_bins': ['TiBins', 'TIBINS', 'tibins', 'TI', 'ti'],
        'sector_weibull': ['SectorWeibull', 'SECTORWEIBULL', 'sectorweibull', 'Sectors', 'SECTORS', 'sectors'],
    }

    def parse(self, file_obj) -> Dict[str, Any]:
        """Parse Excel file."""
        try:
            wb = openpyxl.load_workbook(file_obj, data_only=True)
        except Exception as e:
            raise ParseError(f"Failed to load Excel file: {e}")

        package = {
            'package_version': 'site-package-v1',
            'project': {},
            'site': {},
            'class_envelope': {},
            'layout': {'turbines': []},
            'wtg_models': [],
            'hub_climates': [],
            'gaps': []
        }

        sheets = {name: sheet for name, sheet in zip(wb.sheetnames, wb.worksheets)}

        # Parse each sheet
        self._parse_site_sheet(sheets, package)
        self._parse_layout_sheet(sheets, package)
        self._parse_wtg_models_sheet(sheets, package)
        self._parse_power_curve_sheet(sheets, package)
        self._parse_ct_curve_sheet(sheets, package)
        self._parse_hub_climate_sheet(sheets, package)
        self._parse_ti_bins_sheet(sheets, package)
        self._parse_sector_weibull_sheet(sheets, package)

        return package

    def _find_sheet(self, sheets: Dict[str, Any], sheet_type: str) -> Optional[Any]:
        """Find sheet by aliases."""
        for alias in self.SHEET_ALIASES.get(sheet_type, []):
            if alias in sheets:
                return sheets[alias]
        return None

    def _normalize_header(self, header: str) -> str:
        """Normalize header: lowercase, strip units in [] or ()."""
        if not header:
            return ''
        # Remove units in brackets or parentheses
        import re
        header = re.sub(r'\[.*?\]', '', header)
        header = re.sub(r'\(.*?\)', '', header)
        return header.strip().lower().replace(' ', '_')

    def _parse_site_sheet(self, sheets: Dict[str, Any], package: Dict[str, Any]):
        """Parse Site sheet."""
        sheet = self._find_sheet(sheets, 'site')
        if not sheet:
            return

        # Read key-value pairs (assume 2 columns: key, value)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            key = self._normalize_header(str(row[0])) if row[0] else None
            value = row[1]
            if not key or value is None:
                continue

            if 'name' in key:
                package['site']['name'] = str(value)
                if not package['project'].get('name'):
                    package['project']['name'] = f"{value} Project"
            elif 'lon' in key or 'longitude' in key:
                package['site']['center_lon_deg'] = float(value)
            elif 'lat' in key or 'latitude' in key:
                package['site']['center_lat_deg'] = float(value)
            elif 'complexity' in key:
                complexity = str(value).lower()
                if 'complex' in complexity:
                    package['site']['default_complexity'] = 'complex'
                else:
                    package['site']['default_complexity'] = 'simple'

    def _parse_layout_sheet(self, sheets: Dict[str, Any], package: Dict[str, Any]):
        """Parse Layout/Turbines sheet."""
        sheet = self._find_sheet(sheets, 'layout')
        if not sheet:
            return

        # Read header row
        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
            headers.append(self._normalize_header(str(cell.value)) if cell.value else '')

        # Read turbine rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            turbine = {}
            for idx, value in enumerate(row):
                if idx >= len(headers) or not headers[idx]:
                    continue
                header = headers[idx]

                if 'local_id' in header or 'id' == header:
                    turbine['local_id'] = str(value) if value is not None else None
                elif 'role' in header:
                    role = str(value).lower() if value else 'new_scored'
                    if 'existing' in role or 'wake' in role:
                        turbine['role'] = 'existing_wake_source'
                    else:
                        turbine['role'] = 'new_scored'
                elif 'x_m' in header or header == 'x':
                    turbine['x_m'] = float(value) if value is not None else None
                elif 'y_m' in header or header == 'y':
                    turbine['y_m'] = float(value) if value is not None else None
                elif 'z_base_m' in header or 'z_base' in header or 'elevation' in header:
                    turbine['z_base_m'] = float(value) if value is not None else None
                elif 'hub_height_m' in header or 'hub_height' in header:
                    turbine['hub_height_m'] = float(value) if value is not None else None
                elif 'rotor_d_m' in header or 'rotor_diameter' in header or 'rotor_d' in header:
                    turbine['rotor_d_m'] = float(value) if value is not None else None
                elif 'model_name' in header or 'model' in header:
                    turbine['model_name'] = str(value) if value is not None else None

            if turbine.get('local_id'):
                package['layout']['turbines'].append(turbine)

        if not package['layout'].get('name'):
            package['layout']['name'] = 'Imported Layout'

    def _parse_wtg_models_sheet(self, sheets: Dict[str, Any], package: Dict[str, Any]):
        """Parse WtgModels sheet."""
        sheet = self._find_sheet(sheets, 'wtg_models')
        if not sheet:
            return

        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
            headers.append(self._normalize_header(str(cell.value)) if cell.value else '')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            model = {}
            for idx, value in enumerate(row):
                if idx >= len(headers) or not headers[idx]:
                    continue
                header = headers[idx]

                if 'name' in header:
                    model['name'] = str(value) if value is not None else None
                elif 'rotor_d_m' in header or 'rotor_diameter' in header:
                    model['rotor_d_m'] = float(value) if value is not None else None
                elif 'hub_height' in header:
                    model['hub_height_default_m'] = float(value) if value is not None else None
                elif 'v_in' in header or 'cut_in' in header:
                    model['v_in_mps'] = float(value) if value is not None else None
                elif 'v_rated' in header or 'rated' in header:
                    model['v_rated_mps'] = float(value) if value is not None else None
                elif 'v_out' in header or 'cut_out' in header:
                    model['v_out_mps'] = float(value) if value is not None else None
                elif 'speed_class' in header or header == 'class':
                    # Parse combined class like "IIB" or separate
                    class_str = str(value).upper() if value else ''
                    if len(class_str) >= 2:
                        # Try to split combined class
                        if class_str[-1] in ['A', 'B', 'C', 'S']:
                            model['default_ti_category'] = class_str[-1]
                            speed_part = class_str[:-1]
                            if speed_part in ['I', 'II', 'III', 'S']:
                                model['default_speed_class'] = speed_part
                        elif class_str in ['I', 'II', 'III', 'S']:
                            model['default_speed_class'] = class_str
                elif 'ti_category' in header or 'ti_class' in header:
                    ti = str(value).upper() if value else ''
                    if ti in ['A+', 'A', 'B', 'C', 'S']:
                        model['default_ti_category'] = ti

            if model.get('name'):
                # Initialize empty curves
                model['power_curve'] = []
                model['ct_curve'] = []
                package['wtg_models'].append(model)

    def _parse_power_curve_sheet(self, sheets: Dict[str, Any], package: Dict[str, Any]):
        """Parse PowerCurve sheet."""
        sheet = self._find_sheet(sheets, 'power_curve')
        if not sheet:
            return

        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
            headers.append(self._normalize_header(str(cell.value)) if cell.value else '')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            model_name = None
            v_mps = None
            p_kw = None

            for idx, value in enumerate(row):
                if idx >= len(headers) or not headers[idx]:
                    continue
                header = headers[idx]

                if 'model' in header or 'name' in header:
                    model_name = str(value) if value is not None else None
                elif 'v_mps' in header or header == 'v' or 'speed' in header:
                    v_mps = float(value) if value is not None else None
                elif 'p_kw' in header or 'power' in header:
                    p_kw = float(value) if value is not None else None

            if model_name and v_mps is not None and p_kw is not None:
                # Find model and add point
                for model in package['wtg_models']:
                    if model['name'] == model_name:
                        model['power_curve'].append({'v_mps': v_mps, 'p_kw': p_kw})
                        break

    def _parse_ct_curve_sheet(self, sheets: Dict[str, Any], package: Dict[str, Any]):
        """Parse CtCurve sheet."""
        sheet = self._find_sheet(sheets, 'ct_curve')
        if not sheet:
            return

        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
            headers.append(self._normalize_header(str(cell.value)) if cell.value else '')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            model_name = None
            v_mps = None
            ct = None

            for idx, value in enumerate(row):
                if idx >= len(headers) or not headers[idx]:
                    continue
                header = headers[idx]

                if 'model' in header or 'name' in header:
                    model_name = str(value) if value is not None else None
                elif 'v_mps' in header or header == 'v' or 'speed' in header:
                    v_mps = float(value) if value is not None else None
                elif 'ct' in header or 'thrust' in header:
                    ct = float(value) if value is not None else None

            if model_name and v_mps is not None and ct is not None:
                for model in package['wtg_models']:
                    if model['name'] == model_name:
                        model['ct_curve'].append({'v_mps': v_mps, 'ct': ct})
                        break

    def _parse_hub_climate_sheet(self, sheets: Dict[str, Any], package: Dict[str, Any]):
        """Parse HubClimate sheet."""
        sheet = self._find_sheet(sheets, 'hub_climate')
        if not sheet:
            return

        # Read key-value pairs
        climate = {
            'name': 'Imported Climate',
            'turbine_local_id': None,
            'period_hours': 8760.0,
            'bin_width_mps': 1.0,
            'rho_kgm3': 1.225,
            'ti_bins': [],
            'sector_weibull': []
        }

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 2:
                continue
            key = self._normalize_header(str(row[0])) if row[0] else None
            value = row[1]
            if not key or value is None:
                continue

            if 'name' in key:
                climate['name'] = str(value)
            elif 'period' in key or 'hours' in key:
                climate['period_hours'] = float(value)
            elif 'bin_width' in key:
                climate['bin_width_mps'] = float(value)
            elif 'rho' in key or 'density' in key:
                climate['rho_kgm3'] = float(value)
            elif 'v50' in key or 'v_50' in key:
                climate['v50_mps'] = float(value)
            elif 'shear' in key or 'alpha' in key:
                climate['shear_alpha'] = float(value)
            elif 'inflow' in key or 'angle' in key:
                climate['inflow_angle_deg'] = float(value)

        package['hub_climates'].append(climate)

    def _parse_ti_bins_sheet(self, sheets: Dict[str, Any], package: Dict[str, Any]):
        """Parse TiBins sheet."""
        sheet = self._find_sheet(sheets, 'ti_bins')
        if not sheet:
            return

        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
            headers.append(self._normalize_header(str(cell.value)) if cell.value else '')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            ti_bin = {}
            for idx, value in enumerate(row):
                if idx >= len(headers) or not headers[idx]:
                    continue
                header = headers[idx]

                if 'v_center' in header or 'v_mps' in header or header == 'v':
                    ti_bin['v_center_mps'] = float(value) if value is not None else None
                elif 'hours' in header:
                    ti_bin['hours'] = float(value) if value is not None else None
                elif 'mean_sigma' in header or 'mean' in header:
                    ti_bin['mean_sigma_mps'] = float(value) if value is not None else None
                elif 'std_sigma' in header or 'std' in header:
                    ti_bin['std_sigma_mps'] = float(value) if value not in [None, ''] else None

            if ti_bin.get('v_center_mps') is not None and ti_bin.get('hours') is not None:
                # Add to first climate (assume single climate for Excel)
                if package['hub_climates']:
                    package['hub_climates'][0]['ti_bins'].append(ti_bin)

    def _parse_sector_weibull_sheet(self, sheets: Dict[str, Any], package: Dict[str, Any]):
        """Parse SectorWeibull sheet."""
        sheet = self._find_sheet(sheets, 'sector_weibull')
        if not sheet:
            return

        headers = []
        for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
            headers.append(self._normalize_header(str(cell.value)) if cell.value else '')

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            sector = {}
            for idx, value in enumerate(row):
                if idx >= len(headers) or not headers[idx]:
                    continue
                header = headers[idx]

                if 'from' in header or 'start' in header:
                    sector['sector_from_deg'] = float(value) if value is not None else None
                elif 'to' in header or 'end' in header:
                    sector['sector_to_deg'] = float(value) if value is not None else None
                elif 'frequency' in header or 'freq' in header:
                    sector['frequency'] = float(value) if value is not None else None
                elif header == 'a':
                    sector['A'] = float(value) if value is not None else None
                elif header == 'k':
                    sector['k'] = float(value) if value is not None else None

            if sector.get('sector_from_deg') is not None and sector.get('frequency') is not None:
                if package['hub_climates']:
                    package['hub_climates'][0]['sector_weibull'].append(sector)

        return package


class CSVParser(BaseParser):
    """Parse CSV files."""

    def parse(self, file_obj) -> Dict[str, Any]:
        """Parse CSV file - assume turbines layout."""
        package = {
            'package_version': 'site-package-v1',
            'project': {'name': 'Imported Project'},
            'site': {'name': 'Imported Site', 'center_lon_deg': 0.0, 'center_lat_deg': 0.0, 'default_complexity': 'simple'},
            'class_envelope': {},
            'layout': {'name': 'Imported Layout', 'turbines': []},
            'wtg_models': [],
            'hub_climates': [],
            'gaps': []
        }

        try:
            content = file_obj.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8')

            reader = csv.DictReader(io.StringIO(content))

            for row in reader:
                # Normalize keys
                normalized_row = {k.lower().replace(' ', '_'): v for k, v in row.items() if k}

                turbine = {}
                turbine['local_id'] = normalized_row.get('local_id', normalized_row.get('id'))
                turbine['role'] = normalized_row.get('role', 'new_scored')
                turbine['x_m'] = float(normalized_row['x_m']) if 'x_m' in normalized_row and normalized_row['x_m'] else None
                turbine['y_m'] = float(normalized_row['y_m']) if 'y_m' in normalized_row and normalized_row['y_m'] else None
                turbine['z_base_m'] = float(normalized_row['z_base_m']) if 'z_base_m' in normalized_row and normalized_row['z_base_m'] else None
                turbine['hub_height_m'] = float(normalized_row['hub_height_m']) if 'hub_height_m' in normalized_row and normalized_row['hub_height_m'] else None
                turbine['rotor_d_m'] = float(normalized_row['rotor_d_m']) if 'rotor_d_m' in normalized_row and normalized_row['rotor_d_m'] else None
                turbine['model_name'] = normalized_row.get('model_name')

                if turbine.get('local_id'):
                    package['layout']['turbines'].append(turbine)

        except Exception as e:
            raise ParseError(f"Failed to parse CSV: {e}")

        return package


class HTMLParser(BaseParser):
    """Parse HTML screening reports."""

    def parse(self, file_obj) -> Dict[str, Any]:
        """Parse HTML file."""
        package = {
            'package_version': 'site-package-v1',
            'project': {'name': 'Imported from HTML Report'},
            'site': {'name': 'Imported Site', 'center_lon_deg': 0.0, 'center_lat_deg': 0.0, 'default_complexity': 'simple'},
            'class_envelope': {},
            'layout': {'name': 'Imported Layout', 'turbines': []},
            'wtg_models': [],
            'hub_climates': [],
            'gaps': [{
                'severity': 'run_blocker',
                'path': 'site',
                'code': 'missing_coordinates',
                'message': 'HTML reports often lack coordinates - manual entry required',
                'source_hint': 'HTML parsing limitation'
            }]
        }

        try:
            content = file_obj.read()
            if isinstance(content, bytes):
                content = content.decode('utf-8', errors='ignore')

            soup = BeautifulSoup(content, 'html.parser')

            # Try to extract climate/TI data from tables
            # Note: Do NOT scrape Pass/Fail as climate
            for table in soup.find_all('table'):
                # Look for TI-related headers
                headers = [th.get_text().strip().lower() for th in table.find_all('th')]
                if 'v' in headers or 'speed' in headers or 'sigma' in headers:
                    # Found potential TI table, but skip for now
                    pass

        except Exception as e:
            raise ParseError(f"Failed to parse HTML: {e}")

        return package


def parse_file(file_obj, filename: str) -> Dict[str, Any]:
    """
    Parse uploaded file based on extension.

    Args:
        file_obj: File-like object
        filename: Original filename

    Returns:
        site-package-v1 dict

    Raises:
        ParseError: If file format is not supported or parsing fails
    """
    # Check for rejected formats
    rejected_extensions = ['.map', '.lib', '.shp', '.tif', '.tiff']
    rejected_patterns = ['wasp', 'flowres', 'load_response', 'cfd']

    filename_lower = filename.lower()

    for ext in rejected_extensions:
        if filename_lower.endswith(ext):
            raise ParseError(f"Format {ext} is not supported. Supported formats: JSON, Excel, CSV, HTML.")

    for pattern in rejected_patterns:
        if pattern in filename_lower:
            raise ParseError(f"Format containing '{pattern}' is not supported.")

    # Determine parser based on extension
    if filename_lower.endswith('.json'):
        parser = JSONParser()
    elif filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls'):
        parser = ExcelParser()
    elif filename_lower.endswith('.csv'):
        parser = CSVParser()
    elif filename_lower.endswith('.html') or filename_lower.endswith('.htm'):
        parser = HTMLParser()
    elif filename_lower.endswith('.zip'):
        # Handle ZIP of CSVs
        return _parse_zip(file_obj)
    else:
        raise ParseError(f"Unsupported file format. Supported: JSON, Excel (.xlsx), CSV, HTML, ZIP.")

    return parser.parse(file_obj)


def _parse_zip(file_obj) -> Dict[str, Any]:
    """Parse ZIP file containing multiple CSVs."""
    package = {
        'package_version': 'site-package-v1',
        'project': {'name': 'Imported Project'},
        'site': {},
        'class_envelope': {},
        'layout': {'turbines': []},
        'wtg_models': [],
        'hub_climates': [],
        'gaps': []
    }

    try:
        with zipfile.ZipFile(file_obj, 'r') as zip_ref:
            # Process each CSV file
            for filename in zip_ref.namelist():
                if not filename.lower().endswith('.csv'):
                    continue

                with zip_ref.open(filename) as csv_file:
                    content = csv_file.read().decode('utf-8')
                    reader = csv.DictReader(io.StringIO(content))

                    if 'turbine' in filename.lower() or 'layout' in filename.lower():
                        # Parse turbines
                        for row in reader:
                            normalized_row = {k.lower().replace(' ', '_'): v for k, v in row.items() if k}
                            turbine = {
                                'local_id': normalized_row.get('local_id', normalized_row.get('id')),
                                'role': normalized_row.get('role', 'new_scored'),
                                'x_m': float(normalized_row.get('x_m', 0)) if normalized_row.get('x_m') else None,
                                'y_m': float(normalized_row.get('y_m', 0)) if normalized_row.get('y_m') else None,
                                'z_base_m': float(normalized_row.get('z_base_m', 0)) if normalized_row.get('z_base_m') else None,
                                'hub_height_m': float(normalized_row.get('hub_height_m', 0)) if normalized_row.get('hub_height_m') else None,
                                'rotor_d_m': float(normalized_row.get('rotor_d_m', 0)) if normalized_row.get('rotor_d_m') else None,
                                'model_name': normalized_row.get('model_name'),
                            }
                            if turbine.get('local_id'):
                                package['layout']['turbines'].append(turbine)

    except Exception as e:
        raise ParseError(f"Failed to parse ZIP file: {e}")

    if not package['layout'].get('name'):
        package['layout']['name'] = 'Imported Layout'

    return package
